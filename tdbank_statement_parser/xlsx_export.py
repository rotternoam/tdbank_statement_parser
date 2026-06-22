from collections import defaultdict
from glob import glob
from hashlib import md5
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

from .parser import parse


DEFAULT_COLOR_PALETTE = [
    "FFEB9C",  # light yellow
    "C6EFCE",  # light green
    "FFC7CE",  # light red
    "BDD7EE",  # light blue
    "F8CBAD",  # light orange
    "D9D2E9",  # light purple
    "E2EFDA",  # pale green
    "FCE4D6",  # pale peach
    "DEEAF6",  # pale sky
    "F2DCDB",  # pale rose
]


def build_activity_tables(pdf_paths):
    records = []
    activity_rows = defaultdict(list)

    for filepath in sorted(pdf_paths):
        record = parse(filepath)
        records.append(record)
        for table_name, rows in record["activity"].items():
            for row in rows:
                activity_rows[table_name].append({"filename": record["filename"], **row})

    return records, activity_rows


def build_data_frames(records, activity_rows):
    summary_rows = []
    metadata_rows = []
    transactions_rows = []
    daily_balance_rows = []
    category_rows = defaultdict(lambda: {"count": 0, "total_amount": 0, "transaction_type": None})

    for record in records:
        metadata = record["metadata"]
        beginning_balance = metadata.get("beginning_balance")
        ending_balance = metadata.get("ending_balance")
        total_credits = sum(
            row.get("amount", 0)
            for rows in record["activity"].values()
            for row in rows
            if row.get("transaction_type") == "credit"
        )
        total_debits = sum(
            row.get("amount", 0)
            for rows in record["activity"].values()
            for row in rows
            if row.get("transaction_type") == "debit"
        )
        transaction_count = sum(
            len(rows)
            for table_name, rows in record["activity"].items()
            if table_name != "Daily Balances"
        )

        summary_rows.append(
            {
                "filename": record["filename"],
                "statement_start": metadata.get("statement_period_start"),
                "statement_end": metadata.get("statement_period_end"),
                "beginning_balance": beginning_balance,
                "ending_balance": ending_balance,
                "balance_change": (
                    None
                    if beginning_balance is None or ending_balance is None
                    else ending_balance - beginning_balance
                ),
                "transaction_count": transaction_count,
                "total_credits": total_credits,
                "total_debits": total_debits,
            }
        )

        metadata_rows.append(
            {
                "filename": record["filename"],
                "nPages": record.get("nPages"),
                "statement_period_start": metadata.get("statement_period_start"),
                "statement_period_end": metadata.get("statement_period_end"),
                "customer_reference_number": metadata.get("customer_reference_number"),
                "primary_account_number": metadata.get("primary_account_number"),
                "beginning_balance": beginning_balance,
                "electronic_payments": metadata.get("electronic_payments"),
                "ending_balance": ending_balance,
                "average_collected_balance": metadata.get("average_collected_balance"),
                "interest_earned_this_period": metadata.get("interest_earned_this_period"),
                "interest_paid_ytd": metadata.get("interest_paid_ytd"),
                "annual_percent_yield_earned": metadata.get("annual_percent_yield_earned"),
                "days_in_period": metadata.get("days_in_period"),
                "electronic_deposits": metadata.get("electronic_deposits"),
            }
        )

    for table_name, rows in activity_rows.items():
        if table_name == "Daily Balances":
            for row in rows:
                daily_balance_rows.append(
                    {
                        "filename": row.get("filename"),
                        "date": row.get("date"),
                        "balance": row.get("balance"),
                    }
                )
            continue

        for row in rows:
            transactions_rows.append(
                {
                    "filename": row.get("filename"),
                    "activity_type": table_name,
                    "posting_date": row.get("posting_date"),
                    "description": row.get("description"),
                    "amount": row.get("amount"),
                    "transaction_type": row.get("transaction_type"),
                    "beginning_balance": row.get("beginning_balance"),
                    "ending_balance": row.get("ending_balance"),
                    "daily_balance": row.get("daily_balance"),
                }
            )
            category = category_rows[table_name]
            category["count"] += 1
            category["total_amount"] += row.get("amount", 0) or 0
            if category["transaction_type"] is None:
                category["transaction_type"] = row.get("transaction_type")

    categories_df = pd.DataFrame(
        [
            {
                "activity_type": table_name,
                "count": data["count"],
                "total_amount": data["total_amount"],
                "transaction_type": data["transaction_type"],
            }
            for table_name, data in category_rows.items()
        ]
    )

    return {
        "Summary": pd.DataFrame(summary_rows),
        "Transactions": pd.DataFrame(transactions_rows),
        "Categories": categories_df,
        "Account Metadata": pd.DataFrame(metadata_rows),
        "Daily Balances": pd.DataFrame(daily_balance_rows),
    }


def apply_description_coloring(workbook_path, description_col=4):
    workbook = load_workbook(workbook_path)
    sheet = workbook["Transactions"]
    max_row = sheet.max_row
    max_col = sheet.max_column

    description_colors = {}

    def pastel_color_from_text(text):
        hash_bytes = md5(text.encode("utf-8")).digest()
        r = (hash_bytes[0] + 255) // 2
        g = (hash_bytes[1] + 255) // 2
        b = (hash_bytes[2] + 255) // 2
        return f"{r:02X}{g:02X}{b:02X}"

    for row in range(2, max_row + 1):
        description = sheet.cell(row=row, column=description_col).value
        if description is None:
            continue
        description_key = " ".join(str(description).split())
        if description_key not in description_colors:
            color = pastel_color_from_text(description_key)
            description_colors[description_key] = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid",
            )
        fill = description_colors[description_key]
        for col in range(1, max_col + 1):
            sheet.cell(row=row, column=col).fill = fill

    workbook.save(workbook_path)


def generate_excel(output_path="parsed_statements.xlsx", pdf_folder="inputs"):
    pdf_paths = sorted(glob(f"{pdf_folder}/*.pdf"))
    if not pdf_paths:
        raise FileNotFoundError(f"No PDF files found in {pdf_folder}")

    records, activity_rows = build_activity_tables(pdf_paths)
    sheets = build_data_frames(records, activity_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            if df.empty:
                continue
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    apply_description_coloring(output_path)
    return Path(output_path)


if __name__ == "__main__":
    workbook_path = generate_excel()
    print(f"Generated {workbook_path}")
